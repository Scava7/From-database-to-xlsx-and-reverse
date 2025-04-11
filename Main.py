import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

def scegli_file(titolo, tipi_file):
    file_path = filedialog.askopenfilename(title=titolo, filetypes=tipi_file)
    return file_path

def scegli_salvataggio(titolo, tipo_file):
    file_path = filedialog.asksaveasfilename(title=titolo, defaultextension=tipo_file[1], filetypes=[tipo_file])
    return file_path

def esporta_sqlite_in_excel():
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    import hashlib

    db_path = scegli_file("Scegli il database SQLite", [("Database SQLite", "*.db *.sqlite")])
    if not db_path:
        return

    xlsx_path = scegli_salvataggio("Salva il file Excel", ("Excel Workbook", ".xlsx"))
    if not xlsx_path:
        return

    conn = sqlite3.connect(db_path)
    xls_writer = pd.ExcelWriter(xlsx_path, engine="openpyxl")

    query = "SELECT name FROM sqlite_master WHERE type='table';"
    tables = pd.read_sql(query, conn)

    # Prima scriviamo tutti i fogli
    for table_name in tables['name']:
        df = pd.read_sql(f"SELECT * FROM '{table_name}'", conn)
        df.to_excel(xls_writer, sheet_name=table_name, index=False)

    xls_writer.close()
    conn.close()

    # Poi applichiamo i colori
    wb = load_workbook(xlsx_path)

    for table_name in tables['name']:
        ws = wb[table_name]
        headers = [cell.value for cell in ws[1]]
        if "group" in headers:
            group_index = headers.index("group") + 1  # le colonne partono da 1 in openpyxl
            group_colors = {}


            pastel_colors = [
                "FFCCCC", "FFE5CC", "FFFFCC", "E5FFCC", "CCFFE5",
                "CCFFFF", "CCE5FF", "CCCCFF", "E5CCFF", "FFCCFF",
]
            def get_color(value):
                idx = abs(hash(value)) % len(pastel_colors)
                color = pastel_colors[idx]
                return PatternFill(start_color=color, end_color=color, fill_type="solid")


            for row in ws.iter_rows(min_row=2):
                group_val = row[group_index - 1].value
                if group_val not in group_colors:
                    group_colors[group_val] = get_color(group_val)
                fill = group_colors[group_val]
                for cell in row:
                    cell.fill = fill

    wb.save(xlsx_path)
    messagebox.showinfo("Completato", f"Esportazione completata in:\n{xlsx_path}")


def importa_excel_in_sqlite():
    xlsx_path = scegli_file("Scegli il file Excel", [("Excel Workbook", "*.xlsx")])
    if not xlsx_path:
        return

    db_path = scegli_salvataggio("Salva il database SQLite", ("Database SQLite", ".sqlite"))
    if not db_path:
        return

    xls = pd.ExcelFile(xlsx_path)
    conn = sqlite3.connect(db_path)

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)
        df.to_sql(sheet_name, conn, index=False, if_exists="replace")

    conn.close()
    messagebox.showinfo("Completato", f"Importazione completata in:\n{db_path}")

def main():
    root = tk.Tk()
    root.title("Convertitore SQLite ⇄ Excel")
    root.geometry("400x300")
    root.configure(bg="#f0f4f8")  # azzurrino chiaro

    font_titolo = ("Calibri", 20, "bold")
    font_bottone = ("Calibri", 12)

    titolo = tk.Label(root, text="Convertitore SQLite ⇄ Excel", font=font_titolo, bg="#f0f4f8", fg="#333")
    titolo.pack(pady=(20, 30))

    btn_esporta = tk.Button(root, text="Esporta da SQLite a Excel", font=font_bottone, width=30, command=esporta_sqlite_in_excel)
    btn_esporta.pack(pady=10)

    btn_importa = tk.Button(root, text="Importa da Excel a SQLite", font=font_bottone, width=30, command=importa_excel_in_sqlite)
    btn_importa.pack(pady=10)

    btn_esci = tk.Button(root, text="Esci", font=font_bottone, width=30, command=root.quit)
    btn_esci.pack(pady=20)

    root.mainloop()


if __name__ == "__main__":
    main()
